import math
from dataclasses import dataclass
from typing import Dict, Optional

import pandas as pd
import streamlit as st
import openpyxl


# ============================================================
# Utilities
# ============================================================
def _safe_float(x, default=0.0) -> float:
    try:
        if x is None:
            return float(default)
        if isinstance(x, str) and x.strip().startswith("="):
            # We'll evaluate formulas only if Excel provides cached values; otherwise ignore.
            return float(default)
        return float(x)
    except Exception:
        return float(default)


def _safe_int(x, default=0) -> int:
    try:
        if x is None:
            return int(default)
        if isinstance(x, str) and x.strip().startswith("="):
            return int(default)
        return int(float(x))
    except Exception:
        return int(default)


def los_from_ratio(v_over_c: float) -> str:
    # Same thresholds as your workbook (A<0.35, B<0.55, C<0.75, D<0.86, E<=1, else F)
    if v_over_c < 0.35:
        return "A"
    if v_over_c < 0.55:
        return "B"
    if v_over_c < 0.75:
        return "C"
    if v_over_c < 0.86:
        return "D"
    if v_over_c <= 1.0:
        return "E"
    return "F"


def bpr_time(t0_min: float, v: float, c: float, A: float, B: float) -> float:
    if c <= 0:
        return float("nan")
    x = max(0.0, v / c)
    return t0_min * (1.0 + A * (x ** B))


# ============================================================
# Data model (matches your Inputs + Productivity Loss sheets)
# ============================================================
@dataclass
class Inputs:
    new_office_area_sft: float
    area_per_seat_sft: float
    avg_daily_attendance: float
    share_arriving_busiest_hour: float

    # Mode split (busiest hour)
    share_metro_walk: float
    share_four_wheeler: float
    share_two_wheeler: float
    share_auto: float
    share_bus: float

    # Vehicle occupancies
    occ_four_wheeler: float
    occ_two_wheeler: float
    occ_auto: float
    occ_bus: float

    # PCU factors
    pcu_four_wheeler: float
    pcu_two_wheeler: float
    pcu_auto: float
    pcu_bus: float

    # BPR parameters
    bpr_A: float
    bpr_B: float


@dataclass
class ProductivityParams:
    existing_employees: int
    wfo_days_per_week: int
    working_days_per_month: int

    loss_factor_car: float
    loss_factor_two_wheeler: float
    loss_factor_auto: float
    loss_factor_bus: float

    reliability_buffer_factor: float
    ripple_factor: float

    avg_loaded_annual_salary_inr: float


# ============================================================
# Excel template loader (Cybercity workbook)
# ============================================================
def load_defaults_from_excel(file) -> tuple[Inputs, pd.DataFrame, ProductivityParams]:
    """
    Reads the same cells your workbook uses, so the app starts with your template defaults.
    """
    wb = openpyxl.load_workbook(file, data_only=False)

    # Inputs sheet
    ws = wb["Inputs"]
    inp = Inputs(
        new_office_area_sft=_safe_float(ws["B3"].value, 4800000),
        area_per_seat_sft=_safe_float(ws["B4"].value, 100),
        avg_daily_attendance=_safe_float(ws["B5"].value, 0.65),
        share_arriving_busiest_hour=_safe_float(ws["B6"].value, 0.60),

        share_metro_walk=_safe_float(ws["B13"].value, 0.35),
        share_four_wheeler=_safe_float(ws["C13"].value, 0.30),
        share_two_wheeler=_safe_float(ws["D13"].value, 0.20),
        share_auto=_safe_float(ws["E13"].value, 0.05),
        share_bus=_safe_float(ws["F13"].value, 0.10),

        occ_four_wheeler=_safe_float(ws["B18"].value, 1.5),
        occ_two_wheeler=_safe_float(ws["C18"].value, 1.25),
        occ_auto=_safe_float(ws["D18"].value, 1.5),
        occ_bus=_safe_float(ws["E18"].value, 20),

        pcu_four_wheeler=_safe_float(ws["B23"].value, 1.0),
        pcu_two_wheeler=_safe_float(ws["C23"].value, 0.5),
        pcu_auto=_safe_float(ws["D23"].value, 0.8),
        pcu_bus=_safe_float(ws["E23"].value, 3.0),

        bpr_A=_safe_float(ws["B31"].value, 0.15),
        bpr_B=_safe_float(ws["C31"].value, 4.0),
    )

    # Actual Load (raw present traffic)
    ws_al = wb["Actual Load"]
    # Rows 4-7 are corridors in your template
    corridors = []
    for r in range(4, 8):
        road = ws_al[f"A{r}"].value
        if not road:
            continue
        raw = ws_al[f"C{r}"].value
        # We can't reliably evaluate formulas without cached values; show 0 for formula cells
        raw_val = _safe_float(raw, 0.0)
        veh_hr_formula = ws_al[f"D{r}"].value
        corridors.append({"road": road, "raw_value": raw_val, "vehicles_per_hour_formula": str(veh_hr_formula)})

    # Capacity Vs. Actual: free flow time, capacity, share of added
    ws_c = wb["Capacity Vs. Actual"]
    roads = []
    for r in range(4, 8):
        road = ws_c[f"A{r}"].value
        if not road:
            continue
        t0 = _safe_float(ws_c[f"B{r}"].value, 0.0)
        cap = _safe_float(ws_c[f"C{r}"].value, 0.0)  # may be formula in template
        present = _safe_float(ws_c[f"D{r}"].value, 0.0)  # links to Actual Load
        share = _safe_float(ws_c[f"E{r}"].value, 0.0)
        roads.append(
            {
                "road": road,
                "free_flow_time_min": t0,
                "capacity_pcu_hr": cap,
                "present_flow_pcu_hr": present,
                "share_added_traffic": share,
            }
        )

    roads_df = pd.DataFrame(roads)
    if roads_df.empty:
        # Fallback minimal structure
        roads_df = pd.DataFrame(
            [
                {
                    "road": "Corridor 1",
                    "free_flow_time_min": 10.0,
                    "capacity_pcu_hr": 5000.0,
                    "present_flow_pcu_hr": 4000.0,
                    "share_added_traffic": 0.5,
                },
                {
                    "road": "Corridor 2",
                    "free_flow_time_min": 12.0,
                    "capacity_pcu_hr": 4000.0,
                    "present_flow_pcu_hr": 3000.0,
                    "share_added_traffic": 0.5,
                },
            ]
        )

    # Productivity Loss sheet
    ws_p = wb["Productivity Loss"]
    prod = ProductivityParams(
        existing_employees=_safe_int(ws_p["B4"].value, 500),
        wfo_days_per_week=_safe_int(ws_p["B6"].value, 5),
        working_days_per_month=_safe_int(ws_p["B7"].value, 22),

        loss_factor_car=_safe_float(ws_p["B13"].value, 1.0),
        loss_factor_two_wheeler=_safe_float(ws_p["B14"].value, 1.0),
        loss_factor_auto=_safe_float(ws_p["B15"].value, 0.8),
        loss_factor_bus=_safe_float(ws_p["B16"].value, 0.7),

        reliability_buffer_factor=_safe_float(ws_p["B17"].value, 0.5),
        ripple_factor=_safe_float(ws_p["B18"].value, 0.2),

        avg_loaded_annual_salary_inr=_safe_float(ws_p["B20"].value, 2500000),
    )

    return inp, roads_df, prod


# ============================================================
# Computation (mirrors your workbook flow)
# ============================================================
def validate_mode_shares(inp: Inputs) -> None:
    total = (
        inp.share_metro_walk
        + inp.share_four_wheeler
        + inp.share_two_wheeler
        + inp.share_auto
        + inp.share_bus
    )
    if abs(total - 1.0) > 1e-6:
        raise ValueError(f"Mode shares must sum to 1. Current total = {total:.4f}")


def compute_added_pcu(inp: Inputs) -> Dict[str, float]:
    validate_mode_shares(inp)

    new_seats = inp.new_office_area_sft / inp.area_per_seat_sft if inp.area_per_seat_sft else 0.0
    daily_arrivals = new_seats * inp.avg_daily_attendance
    busiest_hour_people = daily_arrivals * inp.share_arriving_busiest_hour

    # People per road modes (metro/walk -> no vehicles)
    people_four = busiest_hour_people * inp.share_four_wheeler
    people_two = busiest_hour_people * inp.share_two_wheeler
    people_auto = busiest_hour_people * inp.share_auto
    people_bus = busiest_hour_people * inp.share_bus

    veh_four = people_four / inp.occ_four_wheeler if inp.occ_four_wheeler else 0.0
    veh_two = people_two / inp.occ_two_wheeler if inp.occ_two_wheeler else 0.0
    veh_auto = people_auto / inp.occ_auto if inp.occ_auto else 0.0
    veh_bus = people_bus / inp.occ_bus if inp.occ_bus else 0.0

    added_pcu = (
        veh_four * inp.pcu_four_wheeler
        + veh_two * inp.pcu_two_wheeler
        + veh_auto * inp.pcu_auto
        + veh_bus * inp.pcu_bus
    )

    return {
        "new_seats": new_seats,
        "daily_arrivals": daily_arrivals,
        "busiest_hour_people": busiest_hour_people,
        "vehicles_four_wheeler": veh_four,
        "vehicles_two_wheeler": veh_two,
        "vehicles_auto": veh_auto,
        "vehicles_bus": veh_bus,
        "added_pcu": added_pcu,
    }


def compute_roads(inp: Inputs, roads_df: pd.DataFrame, added_pcu_total: float) -> pd.DataFrame:
    df = roads_df.copy()

    required_cols = {
        "road", "free_flow_time_min", "capacity_pcu_hr", "present_flow_pcu_hr", "share_added_traffic"
    }
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(f"Roads table is missing columns: {sorted(missing)}")

    share_sum = float(df["share_added_traffic"].sum())
    if share_sum <= 0:
        raise ValueError("Sum of 'share_added_traffic' must be > 0.")

    df["share_added_traffic_norm"] = df["share_added_traffic"] / share_sum
    df["added_pcu"] = added_pcu_total * df["share_added_traffic_norm"]
    df["future_flow_pcu_hr"] = df["present_flow_pcu_hr"] + df["added_pcu"]

    df["vc_now"] = df["present_flow_pcu_hr"] / df["capacity_pcu_hr"]
    df["vc_future"] = df["future_flow_pcu_hr"] / df["capacity_pcu_hr"]

    df["time_now_min"] = df.apply(
        lambda r: bpr_time(r["free_flow_time_min"], r["present_flow_pcu_hr"], r["capacity_pcu_hr"], inp.bpr_A, inp.bpr_B),
        axis=1,
    )
    df["time_future_min"] = df.apply(
        lambda r: bpr_time(r["free_flow_time_min"], r["future_flow_pcu_hr"], r["capacity_pcu_hr"], inp.bpr_A, inp.bpr_B),
        axis=1,
    )
    df["extra_min"] = df["time_future_min"] - df["time_now_min"]

    df["los_now"] = df["vc_now"].apply(los_from_ratio)
    df["los_future"] = df["vc_future"].apply(los_from_ratio)

    return df[
        [
            "road",
            "free_flow_time_min",
            "capacity_pcu_hr",
            "present_flow_pcu_hr",
            "added_pcu",
            "future_flow_pcu_hr",
            "vc_now",
            "vc_future",
            "time_now_min",
            "time_future_min",
            "extra_min",
            "los_now",
            "los_future",
        ]
    ]


def compute_final_outcome(inp: Inputs, roads_out: pd.DataFrame, added: Dict[str, float]) -> Dict[str, float]:
    avg_now = float(roads_out["time_now_min"].mean())
    avg_future = float(roads_out["time_future_min"].mean())
    avg_extra = avg_future - avg_now

    roads_at_jam_future = int((roads_out["los_future"].isin(["E", "F"])).sum())

    share_using_roads_people = 1.0 - inp.share_metro_walk
    daily_team_hours_lost_new_supply = (added["daily_arrivals"] * avg_extra / 60.0) * share_using_roads_people

    return {
        "avg_last_mile_now_min": avg_now,
        "avg_last_mile_future_min": avg_future,
        "avg_extra_min": avg_extra,
        "roads_at_jam_future": roads_at_jam_future,
        "share_using_roads_people": share_using_roads_people,
        "daily_team_hours_lost_new_supply": daily_team_hours_lost_new_supply,
    }


def compute_productivity_loss(inp: Inputs, prod: ProductivityParams, roads_out: pd.DataFrame) -> Dict[str, float]:
    # Weighting corridors using added_pcu distribution (same logic as the workbook’s corridor focus)
    total_added = float(roads_out["added_pcu"].sum())
    if total_added > 0:
        weights = roads_out["added_pcu"] / total_added
    else:
        weights = pd.Series([1.0 / len(roads_out)] * len(roads_out))

    avg_extra_min_existing = float((roads_out["extra_min"] * weights).sum())

    road_share = inp.share_four_wheeler + inp.share_two_wheeler + inp.share_auto + inp.share_bus
    if road_share <= 0:
        raise ValueError("Road share (non-metro) must be > 0.")

    arrivals_per_day_in_office = prod.existing_employees * inp.avg_daily_attendance * (prod.wfo_days_per_week / 5.0)
    road_arrivals_per_day = arrivals_per_day_in_office * road_share

    weighted_loss_factor = (
        (inp.share_four_wheeler * prod.loss_factor_car)
        + (inp.share_two_wheeler * prod.loss_factor_two_wheeler)
        + (inp.share_auto * prod.loss_factor_auto)
        + (inp.share_bus * prod.loss_factor_bus)
    ) / road_share

    base_hours = road_arrivals_per_day * avg_extra_min_existing / 60.0
    direct_hours_lost_per_day = base_hours * weighted_loss_factor
    reliability_buffer_hours_per_day = base_hours * prod.reliability_buffer_factor
    ripple_hours_per_day = direct_hours_lost_per_day * prod.ripple_factor
    total_hours_lost_per_day = direct_hours_lost_per_day + reliability_buffer_hours_per_day + ripple_hours_per_day

    effective_in_office_days_month = prod.working_days_per_month * (prod.wfo_days_per_week / 5.0)
    total_hours_lost_per_month = total_hours_lost_per_day * effective_in_office_days_month

    avg_monthly_salary = prod.avg_loaded_annual_salary_inr / 12.0
    inr_per_hour = avg_monthly_salary / prod.working_days_per_month / 9.0

    cost_per_day = total_hours_lost_per_day * inr_per_hour
    cost_per_month = total_hours_lost_per_month * inr_per_hour

    return {
        "avg_extra_min_existing": avg_extra_min_existing,
        "arrivals_per_day_in_office": arrivals_per_day_in_office,
        "road_arrivals_per_day": road_arrivals_per_day,
        "weighted_loss_factor": weighted_loss_factor,
        "direct_hours_lost_per_day": direct_hours_lost_per_day,
        "reliability_buffer_hours_per_day": reliability_buffer_hours_per_day,
        "ripple_hours_per_day": ripple_hours_per_day,
        "total_hours_lost_per_day": total_hours_lost_per_day,
        "total_hours_lost_per_month": total_hours_lost_per_month,
        "inr_per_hour": inr_per_hour,
        "cost_per_day_inr": cost_per_day,
        "cost_per_month_inr": cost_per_month,
    }


# ============================================================
# Streamlit App
# ============================================================
st.set_page_config(page_title="Micromarket Supply Impact App", layout="wide")
st.title("Micromarket Supply Impact App (Traffic + Productivity Loss)")
st.caption("Upload your Excel template (optional) or edit inputs directly. Output = extra travel time + productivity loss for existing occupiers.")


with st.sidebar:
    st.header("Start from Excel template")
    uploaded = st.file_uploader("Upload your Cybercity model .xlsx", type=["xlsx"])
    st.write("If you upload the workbook, the app pre-fills the inputs from the same cells you shared (Inputs, Capacity Vs. Actual, Productivity Loss).")

# Defaults
if "defaults_loaded" not in st.session_state:
    st.session_state.defaults_loaded = False

if uploaded and not st.session_state.defaults_loaded:
    try:
        inp0, roads0, prod0 = load_defaults_from_excel(uploaded)
        st.session_state.inp0 = inp0
        st.session_state.roads0 = roads0
        st.session_state.prod0 = prod0
        st.session_state.defaults_loaded = True
        st.success("Loaded defaults from the uploaded Excel template.")
    except Exception as e:
        st.warning(f"Could not load all defaults from Excel. Using fallback defaults. Details: {e}")

if st.session_state.defaults_loaded:
    inp = st.session_state.inp0
    roads_df = st.session_state.roads0
    prod = st.session_state.prod0
else:
    # Sensible fallback aligned to your Cybercity sheet
    inp = Inputs(
        new_office_area_sft=4_800_000,
        area_per_seat_sft=100.0,
        avg_daily_attendance=0.65,
        share_arriving_busiest_hour=0.60,
        share_metro_walk=0.35,
        share_four_wheeler=0.30,
        share_two_wheeler=0.20,
        share_auto=0.05,
        share_bus=0.10,
        occ_four_wheeler=1.5,
        occ_two_wheeler=1.25,
        occ_auto=1.5,
        occ_bus=20.0,
        pcu_four_wheeler=1.0,
        pcu_two_wheeler=0.5,
        pcu_auto=0.8,
        pcu_bus=3.0,
        bpr_A=0.15,
        bpr_B=4.0,
    )
    roads_df = pd.DataFrame(
        [
            {"road": "Dhaula Kuan to NH-48 to Cyber City", "free_flow_time_min": 16.0, "capacity_pcu_hr": 9200.0, "present_flow_pcu_hr": 16000.0, "share_added_traffic": 0.30},
            {"road": "Chattarpur to Cyber City", "free_flow_time_min": 14.0, "capacity_pcu_hr": 6900.0, "present_flow_pcu_hr": 0.0, "share_added_traffic": 0.25},
            {"road": "Golf Course Road to Cyber City", "free_flow_time_min": 10.0, "capacity_pcu_hr": 3600.0, "present_flow_pcu_hr": 0.0, "share_added_traffic": 0.20},
            {"road": "Rajiv Chowk to Cyber City", "free_flow_time_min": 8.0, "capacity_pcu_hr": 9200.0, "present_flow_pcu_hr": 16000.0, "share_added_traffic": 0.25},
        ]
    )
    prod = ProductivityParams(
        existing_employees=500,
        wfo_days_per_week=5,
        working_days_per_month=22,
        loss_factor_car=1.0,
        loss_factor_two_wheeler=1.0,
        loss_factor_auto=0.8,
        loss_factor_bus=0.7,
        reliability_buffer_factor=0.5,
        ripple_factor=0.2,
        avg_loaded_annual_salary_inr=2_500_000.0,
    )

# Layout
c1, c2 = st.columns([1.1, 0.9], gap="large")

with c1:
    st.subheader("Inputs (blue cells)")
    inp.new_office_area_sft = st.number_input("New office area (sft)", min_value=0.0, value=float(inp.new_office_area_sft), step=50_000.0)
    inp.area_per_seat_sft = st.number_input("Area per seat (sft)", min_value=10.0, value=float(inp.area_per_seat_sft), step=5.0)
    inp.avg_daily_attendance = st.number_input("Average daily attendance", min_value=0.0, max_value=1.0, value=float(inp.avg_daily_attendance), step=0.01)
    inp.share_arriving_busiest_hour = st.number_input("Share of arriving people in the busiest hour", min_value=0.0, max_value=1.0, value=float(inp.share_arriving_busiest_hour), step=0.01)

    st.markdown("### Mode split (must sum to 1)")
    inp.share_metro_walk = st.number_input("Metro/Walk", min_value=0.0, max_value=1.0, value=float(inp.share_metro_walk), step=0.01)
    inp.share_four_wheeler = st.number_input("Four‑wheeler", min_value=0.0, max_value=1.0, value=float(inp.share_four_wheeler), step=0.01)
    inp.share_two_wheeler = st.number_input("Two‑wheeler", min_value=0.0, max_value=1.0, value=float(inp.share_two_wheeler), step=0.01)
    inp.share_auto = st.number_input("Auto‑rickshaw", min_value=0.0, max_value=1.0, value=float(inp.share_auto), step=0.01)
    inp.share_bus = st.number_input("Bus/Shuttle", min_value=0.0, max_value=1.0, value=float(inp.share_bus), step=0.01)

    st.markdown("### Vehicle occupancies")
    inp.occ_four_wheeler = st.number_input("Four‑wheeler occupancy", min_value=0.1, value=float(inp.occ_four_wheeler), step=0.05)
    inp.occ_two_wheeler = st.number_input("Two‑wheeler occupancy", min_value=0.1, value=float(inp.occ_two_wheeler), step=0.05)
    inp.occ_auto = st.number_input("Auto occupancy", min_value=0.1, value=float(inp.occ_auto), step=0.05)
    inp.occ_bus = st.number_input("Bus occupancy", min_value=1.0, value=float(inp.occ_bus), step=1.0)

    st.markdown("### PCU factors")
    inp.pcu_four_wheeler = st.number_input("PCU – Four‑wheeler", min_value=0.0, value=float(inp.pcu_four_wheeler), step=0.1)
    inp.pcu_two_wheeler = st.number_input("PCU – Two‑wheeler", min_value=0.0, value=float(inp.pcu_two_wheeler), step=0.1)
    inp.pcu_auto = st.number_input("PCU – Auto", min_value=0.0, value=float(inp.pcu_auto), step=0.1)
    inp.pcu_bus = st.number_input("PCU – Bus", min_value=0.0, value=float(inp.pcu_bus), step=0.1)

    st.markdown("### Delay curve (BPR)")
    inp.bpr_A = st.number_input("A", min_value=0.0, value=float(inp.bpr_A), step=0.01)
    inp.bpr_B = st.number_input("B", min_value=0.1, value=float(inp.bpr_B), step=0.1)

with c2:
    st.subheader("Corridors (Actual Load + Capacity Vs. Actual)")
    st.write("Edit corridors: free-flow time, capacity, present flow, and share of added traffic.")

    roads_df = st.data_editor(
        roads_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "free_flow_time_min": st.column_config.NumberColumn(format="%.2f"),
            "capacity_pcu_hr": st.column_config.NumberColumn(format="%.0f"),
            "present_flow_pcu_hr": st.column_config.NumberColumn(format="%.0f"),
            "share_added_traffic": st.column_config.NumberColumn(format="%.3f"),
        },
    )

    st.divider()
    st.subheader("Existing occupier productivity parameters")
    prod.existing_employees = st.number_input("Existing employees (seats)", min_value=0, value=int(prod.existing_employees), step=50)
    prod.wfo_days_per_week = st.number_input("WFO days/week", min_value=0, max_value=7, value=int(prod.wfo_days_per_week), step=1)
    prod.working_days_per_month = st.number_input("Working days/month", min_value=1, max_value=31, value=int(prod.working_days_per_month), step=1)

    prod.loss_factor_car = st.number_input("Loss factor – Car driver", min_value=0.0, max_value=1.0, value=float(prod.loss_factor_car), step=0.05)
    prod.loss_factor_two_wheeler = st.number_input("Loss factor – Two‑wheeler", min_value=0.0, max_value=1.0, value=float(prod.loss_factor_two_wheeler), step=0.05)
    prod.loss_factor_auto = st.number_input("Loss factor – Auto", min_value=0.0, max_value=1.0, value=float(prod.loss_factor_auto), step=0.05)
    prod.loss_factor_bus = st.number_input("Loss factor – Bus/Shuttle", min_value=0.0, max_value=1.0, value=float(prod.loss_factor_bus), step=0.05)

    prod.reliability_buffer_factor = st.number_input("Reliability buffer factor", min_value=0.0, max_value=2.0, value=float(prod.reliability_buffer_factor), step=0.05)
    prod.ripple_factor = st.number_input("Ripple factor", min_value=0.0, max_value=2.0, value=float(prod.ripple_factor), step=0.05)

    prod.avg_loaded_annual_salary_inr = st.number_input("Avg fully loaded annual salary (₹)", min_value=0.0, value=float(prod.avg_loaded_annual_salary_inr), step=50_000.0)

st.divider()

# Run calculations
try:
    added = compute_added_pcu(inp)
    roads_out = compute_roads(inp, roads_df, added["added_pcu"])
    final = compute_final_outcome(inp, roads_out, added)
    prod_out = compute_productivity_loss(inp, prod, roads_out)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Avg extra last‑mile time", f"{final['avg_extra_min']:.2f} min")
    k2.metric("Corridors at LOS E/F (future)", f"{final['roads_at_jam_future']}")
    k3.metric("Hours lost/day (new supply arrivals)", f"{final['daily_team_hours_lost_new_supply']:.2f} hrs/day")
    k4.metric("Productivity loss/month", f"₹{prod_out['cost_per_month_inr']:.0f}")

    st.subheader("Road-wise output (Capacity Vs. Actual)")
    st.dataframe(roads_out, use_container_width=True)

    st.subheader("Productivity loss (existing occupiers)")
    p1, p2, p3, p4 = st.columns(4)
    p1.metric("Avg extra mins (existing, weighted)", f"{prod_out['avg_extra_min_existing']:.2f} min")
    p2.metric("Total hours lost/day", f"{prod_out['total_hours_lost_per_day']:.2f} hrs")
    p3.metric("Cost/day", f"₹{prod_out['cost_per_day_inr']:.0f}")
    p4.metric("Cost/month", f"₹{prod_out['cost_per_month_inr']:.0f}")

    with st.expander("Show calculation breakdown (for audit)"):
        st.json({"added_supply": added, "final_outcome": final, "productivity": prod_out})

except Exception as e:
    st.error(f"Model error: {e}")
